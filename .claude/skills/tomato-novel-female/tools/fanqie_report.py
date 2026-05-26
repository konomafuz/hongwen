#!/usr/bin/env python3
"""番茄小说女频排行榜数据报告生成器

读取 fanqie_scraper.py 输出的 JSON 数据，生成 Excel 或 CSV 报告。

使用方法：
    python fanqie_report.py data.json                      # 默认输出 Excel
    python fanqie_report.py data.json -o report.xlsx       # 指定输出路径
    python fanqie_report.py data.json --format csv         # 输出 CSV（无需 openpyxl）

依赖：openpyxl（可选，无此库时自动降级为 CSV）
      pip install openpyxl
"""

import argparse
import csv
import json
import os
import re
import sys
from collections import Counter
from datetime import datetime
from itertools import combinations

# ── 停用词表（用于从简介中提取关键词时过滤）──────────

STOPWORDS = set(
    "的 了 在 是 我 有 和 就 不 人 都 一 一个 上 也 很 到 说 要 去 你 会 着 没有 看 好 "
    "自己 这 他 她 它 们 那 里 个 么 什么 吗 吧 啊 呢 把 被 从 让 给 对 又 但 还 可以 "
    "能 为 与 以 及 之 而 或 等 却 已 已经 虽然 因为 所以 如果 就是 这个 那个 的话 "
    "可是 并 却 只 只是 更 最 过 将 得 地 来 去 做 能够 不是 没 时候 什么样 怎么 "
    "一种 一些 这些 那些 每 各 两 第 所有 其他 其中 之后 之前 然后 不过 只要 "
    "无论 任何 通过 关于 为了 除了 以及 或者 而是 不仅 而且 同时 但是 然而 "
    "因此 于是 终于 居然 竟然 原来 居然 简直 几乎 完全 真的 确实 其实 当然".split()
)


# ── 数据分析函数 ──────────────────────────────────────

def analyze_tags(books: list[dict]) -> list[dict]:
    """统计标签频率并找出常见搭配。"""
    tag_counter = Counter()
    tag_cooccur = Counter()

    for book in books:
        tags = book.get("tags", [])
        for tag in tags:
            tag_counter[tag] += 1
        # 标签两两共现
        for pair in combinations(sorted(set(tags)), 2):
            tag_cooccur[pair] += 1

    results = []
    for tag, count in tag_counter.most_common(20):
        # 找出最常见的搭配标签
        co_tags = []
        for (a, b), co_count in tag_cooccur.most_common():
            if a == tag:
                co_tags.append(b)
            elif b == tag:
                co_tags.append(a)
            if len(co_tags) >= 3:
                break
        results.append({
            "标签": tag,
            "频次": count,
            "占比": f"{count / len(books) * 100:.1f}%",
            "常见搭配": "、".join(co_tags) if co_tags else "-",
        })

    return results


def analyze_categories(books: list[dict]) -> list[dict]:
    """统计题材（分类）分布。"""
    cat_counter = Counter()
    cat_examples = {}

    for book in books:
        cat = book.get("category", "未知")
        cat_counter[cat] += 1
        if cat not in cat_examples:
            cat_examples[cat] = []
        if len(cat_examples[cat]) < 3:
            cat_examples[cat].append(book.get("title", ""))

    results = []
    for cat, count in cat_counter.most_common(10):
        results.append({
            "题材": cat,
            "数量": count,
            "占比": f"{count / len(books) * 100:.1f}%",
            "代表作": "、".join(cat_examples.get(cat, [])[:3]),
        })

    return results


def extract_keywords(books: list[dict], top_n: int = 30) -> list[dict]:
    """从简介和书名中提取高频关键词。"""
    word_counter = Counter()

    for book in books:
        text = (book.get("title", "") + " " + book.get("abstract", ""))
        # 简单分词：提取2-4字的中文词组
        words = re.findall(r"[\u4e00-\u9fff]{2,4}", text)
        for w in words:
            if w not in STOPWORDS and len(w) >= 2:
                word_counter[w] += 1

    results = []
    for word, count in word_counter.most_common(top_n):
        results.append({
            "热词": word,
            "频次": count,
        })

    return results


def analyze_character_types(books: list[dict]) -> dict:
    """从简介中识别常见人设类型。"""
    # 女主人设关键词
    female_lead_patterns = {
        "重生女强": ["重生", "浴火重生", "重活一世", "回到过去"],
        "穿书女配": ["穿书", "穿进", "书中", "女配"],
        "替嫁新娘": ["替嫁", "代嫁", "冲喜"],
        "天才女主": ["天才", "废材逆袭", "修炼奇才", "医术超群"],
        "隐忍复仇": ["复仇", "报仇", "雪恨", "隐忍"],
        "甜宠娇妻": ["甜宠", "宠妻", "娇妻", "溺爱", "团宠"],
        "飒爽女王": ["女王", "女强", "霸气", "飒"],
        "落魄千金": ["落魄", "千金", "豪门", "家道中落"],
        "软萌学霸": ["学霸", "软萌", "呆萌", "可爱"],
        "职场精英": ["总裁", "职场", "精英", "白领"],
    }

    # 男主人设关键词
    male_lead_patterns = {
        "冷面总裁": ["冷面", "冰山", "高冷", "总裁"],
        "忠犬暖男": ["忠犬", "暖男", "温柔", "深情"],
        "腹黑权臣": ["腹黑", "权臣", "城府", "心机"],
        "病娇偏执": ["病娇", "偏执", "疯批", "黑化"],
        "战神王爷": ["战神", "王爷", "将军", "元帅"],
        "禁欲系": ["禁欲", "清冷", "仙君", "淡漠"],
        "痞帅少年": ["痞帅", "少年", "校草", "学长"],
        "权势滔天": ["权势", "首富", "家主", "掌权"],
    }

    female_counter = Counter()
    male_counter = Counter()

    for book in books:
        text = book.get("abstract", "") + " " + " ".join(book.get("tags", []))
        for char_type, keywords in female_lead_patterns.items():
            if any(kw in text for kw in keywords):
                female_counter[char_type] += 1
        for char_type, keywords in male_lead_patterns.items():
            if any(kw in text for kw in keywords):
                male_counter[char_type] += 1

    female_results = [
        {"人设类型": t, "频次": c}
        for t, c in female_counter.most_common(5)
    ]
    male_results = [
        {"人设类型": t, "频次": c}
        for t, c in male_counter.most_common(5)
    ]

    return {"女主人设TOP5": female_results, "男主人设TOP5": male_results}


def analyze_tag_combos(books: list[dict], top_n: int = 10) -> list[dict]:
    """分析标签组合（2-3标签共现）。"""
    combo2_counter = Counter()
    combo_examples = {}

    for book in books:
        tags = sorted(set(book.get("tags", [])))
        title = book.get("title", "")
        for combo in combinations(tags, 2):
            combo2_counter[combo] += 1
            key = " + ".join(combo)
            if key not in combo_examples:
                combo_examples[key] = []
            if len(combo_examples[key]) < 2:
                combo_examples[key].append(title)

    results = []
    for combo, count in combo2_counter.most_common(top_n):
        key = " + ".join(combo)
        results.append({
            "标签组合": key,
            "出现次数": count,
            "代表作": "、".join(combo_examples.get(key, [])),
        })

    return results


# ── 报告输出 ──────────────────────────────────────────

def write_excel(
    output_path: str,
    books: list[dict],
    categories: list[dict],
    tags: list[dict],
    keywords: list[dict],
    characters: dict,
    combos: list[dict],
) -> None:
    """输出 Excel 报告（需要 openpyxl）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def write_sheet(ws, title: str, data: list[dict]) -> None:
        """写入一个工作表。"""
        ws.title = title
        if not data:
            ws.append(["暂无数据"])
            return

        # 表头
        headers = list(data[0].keys())
        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 数据行
        for row_data in data:
            row = [row_data.get(h, "") for h in headers]
            ws.append(row)

        # 自动列宽（简单估算）
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row in data:
                val = str(row.get(header, ""))
                max_len = max(max_len, min(len(val), 50))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 4

    # Sheet 1: 热门题材
    ws1 = wb.active
    write_sheet(ws1, "热门题材TOP10", categories)

    # Sheet 2: 热门标签
    ws2 = wb.create_sheet()
    write_sheet(ws2, "热门标签TOP20", tags)

    # Sheet 3: 热词
    ws3 = wb.create_sheet()
    write_sheet(ws3, "热词TOP30", keywords)

    # Sheet 4: 热门人设
    ws4 = wb.create_sheet()
    ws4.title = "热门人设"
    # 女主
    ws4.append(["== 女主人设 TOP5 =="])
    ws4.cell(row=1, column=1).font = Font(bold=True, size=12)
    if characters.get("女主人设TOP5"):
        headers = list(characters["女主人设TOP5"][0].keys())
        ws4.append(headers)
        for cell_idx in range(1, len(headers) + 1):
            ws4.cell(row=2, column=cell_idx).font = header_font
        for item in characters["女主人设TOP5"]:
            ws4.append([item.get(h, "") for h in headers])
    # 男主
    gap_row = ws4.max_row + 2
    ws4.cell(row=gap_row, column=1, value="== 男主人设 TOP5 ==").font = Font(bold=True, size=12)
    if characters.get("男主人设TOP5"):
        headers = list(characters["男主人设TOP5"][0].keys())
        ws4.append(headers)
        row_num = ws4.max_row
        for cell_idx in range(1, len(headers) + 1):
            ws4.cell(row=row_num, column=cell_idx).font = header_font
        for item in characters["男主人设TOP5"]:
            ws4.append([item.get(h, "") for h in headers])

    # Sheet 5: 标签组合
    ws5 = wb.create_sheet()
    write_sheet(ws5, "标签组合TOP10", combos)

    # Sheet 6: 原始数据
    ws6 = wb.create_sheet()
    raw_data = []
    for book in books:
        raw_data.append({
            "书名": book.get("title", ""),
            "作者": book.get("author", ""),
            "分类": book.get("category", ""),
            "标签": "、".join(book.get("tags", [])),
            "字数": book.get("word_count", 0),
            "阅读量": book.get("read_count", 0),
            "简介": book.get("abstract", "")[:200],
            "榜单": book.get("rank_type", ""),
            "book_id": book.get("book_id", ""),
        })
    write_sheet(ws6, "原始数据", raw_data)

    wb.save(output_path)


def write_csv(output_dir: str, books: list[dict], categories, tags, keywords, characters, combos):
    """输出多个 CSV 文件（openpyxl 不可用时的降级方案）。"""
    os.makedirs(output_dir, exist_ok=True)

    def save_csv(filename: str, data: list[dict]) -> None:
        if not data:
            return
        path = os.path.join(output_dir, filename)
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        print(f"  已生成 {path}", file=sys.stderr)

    save_csv("01_热门题材TOP10.csv", categories)
    save_csv("02_热门标签TOP20.csv", tags)
    save_csv("03_热词TOP30.csv", keywords)
    save_csv("04_女主人设TOP5.csv", characters.get("女主人设TOP5", []))
    save_csv("05_男主人设TOP5.csv", characters.get("男主人设TOP5", []))
    save_csv("06_标签组合TOP10.csv", combos)

    # 原始数据
    raw_data = []
    for book in books:
        raw_data.append({
            "书名": book.get("title", ""),
            "作者": book.get("author", ""),
            "分类": book.get("category", ""),
            "标签": "、".join(book.get("tags", [])),
            "字数": book.get("word_count", 0),
            "阅读量": book.get("read_count", 0),
            "简介": book.get("abstract", "")[:200],
            "榜单": book.get("rank_type", ""),
        })
    save_csv("07_原始数据.csv", raw_data)


# ── 主流程 ──────────────────────────────────────────

def generate_report(data: dict, output_path: str, output_format: str = "auto") -> None:
    """生成报告。"""
    books = data.get("books", [])
    if not books:
        print("错误：JSON 数据中没有书籍记录", file=sys.stderr)
        sys.exit(1)

    print(f"分析 {len(books)} 本书的数据...", file=sys.stderr)

    # 执行分析
    categories = analyze_categories(books)
    tags = analyze_tags(books)
    keywords = extract_keywords(books)
    characters = analyze_character_types(books)
    combos = analyze_tag_combos(books)

    # 决定输出格式
    if output_format == "auto":
        try:
            import openpyxl
            output_format = "excel"
        except ImportError:
            print("提示：未安装 openpyxl，将输出 CSV 格式。安装方法：pip install openpyxl", file=sys.stderr)
            output_format = "csv"

    if output_format == "excel":
        if not output_path.endswith(".xlsx"):
            output_path = output_path + ".xlsx" if not output_path.endswith(".csv") else output_path.replace(".csv", ".xlsx")
        write_excel(output_path, books, categories, tags, keywords, characters, combos)
        print(f"Excel 报告已生成：{output_path}", file=sys.stderr)
    else:
        # CSV 模式：output_path 作为目录
        csv_dir = output_path if not output_path.endswith((".xlsx", ".csv")) else os.path.splitext(output_path)[0]
        write_csv(csv_dir, books, categories, tags, keywords, characters, combos)
        print(f"CSV 报告已生成到目录：{csv_dir}/", file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(
        description="番茄小说女频排行榜数据报告生成器",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="示例：\n"
               "  python fanqie_report.py data.json\n"
               "  python fanqie_report.py data.json -o report.xlsx\n"
               "  python fanqie_report.py data.json --format csv -o ./reports\n",
    )
    parser.add_argument("input", help="爬虫输出的 JSON 文件路径")
    parser.add_argument("-o", "--output", default="fanqie_report", help="输出路径（默认 fanqie_report）")
    parser.add_argument(
        "--format",
        choices=["auto", "excel", "csv"],
        default="auto",
        help="输出格式（默认 auto：有 openpyxl 则 Excel，否则 CSV）",
    )

    args = parser.parse_args()

    # 读取 JSON
    try:
        with open(args.input, "r", encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"错误：文件不存在 {args.input}", file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"错误：JSON 解析失败 {e}", file=sys.stderr)
        sys.exit(1)

    generate_report(data, args.output, args.format)


if __name__ == "__main__":
    main()
